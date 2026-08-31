"""Library export routes for the DiscVault Next backend.

Exposes the column catalogue and renders a Library selection to CSV, XLSX or
PDF. The frontend posts the rows exactly as they are displayed in the list view
(see ``next_export_columns`` for why), so an export is always WYSIWYG: it keeps
the active filters, the active sort order and one row per film.
"""

from __future__ import annotations

import csv
import io
import re
from datetime import datetime, timezone
from typing import Any

from flask import Flask, request

try:  # pragma: no cover - exercised indirectly by both layouts
    from .next_common import NextApiError, response, table_exists
    from .next_export_columns import (
        column_headers,
        column_weights,
        export_column_catalogue,
        normalize_columns,
        normalize_rows,
    )
except ImportError:  # pragma: no cover - supports gunicorn next_app:app
    from next_common import NextApiError, response, table_exists
    from next_export_columns import (
        column_headers,
        column_weights,
        export_column_catalogue,
        normalize_columns,
        normalize_rows,
    )


EXPORT_FORMATS = ("csv", "xlsx", "pdf")

EXPORT_MIME_TYPES = {
    "csv": "text/csv; charset=utf-8",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf": "application/pdf",
}

DEFAULT_EXPORT_TITLE = "DiscVault library"

# ReportLab builds one Paragraph per cell and lays the whole table out in memory,
# synchronously, inside the request. The deployment runs two gunicorn workers, so an
# unbounded PDF render occupies half the app's request capacity for its full duration
# and can be killed by the worker timeout mid-flight. CSV and XLSX stream far more
# cheaply and carry no such ceiling.
MAX_PDF_EXPORT_ROWS = 5_000

_FILENAME_SAFE = re.compile(r"[^A-Za-z0-9._-]+")


def _next_app():
    try:  # pragma: no cover - import shape depends on runtime layout
        from . import next_app
    except ImportError:  # pragma: no cover - supports gunicorn next_app:app
        import next_app
    return next_app


def normalize_format(value: Any) -> str:
    fmt = str(value or "csv").strip().lower()
    if fmt not in EXPORT_FORMATS:
        raise NextApiError(f"Unsupported export format: {fmt}", 400)
    return fmt


def export_filename(fmt: str, requested: Any = None, *, now: datetime | None = None) -> str:
    """Build a safe ``discvault-library-<date>.<ext>`` style filename."""
    stem = _FILENAME_SAFE.sub("-", str(requested or "").strip()).strip("-._")
    if stem.lower().endswith(f".{fmt}"):
        stem = stem[: -(len(fmt) + 1)]
    if not stem:
        stamp = (now or datetime.now(timezone.utc)).strftime("%Y%m%d")
        stem = f"discvault-library-{stamp}"
    return f"{stem[:120]}.{fmt}"


def render_csv(headers: list[str], rows: list[list[str]]) -> bytes:
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer, delimiter=";", quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")
    writer.writerow(headers)
    for row in rows:
        writer.writerow([value.replace("\n", " / ") for value in row])
    # A BOM keeps Excel from mangling accented titles on a double click.
    return b"\xef\xbb\xbf" + buffer.getvalue().encode("utf-8")


def render_xlsx(headers: list[str], rows: list[list[str]], *, sheet_title: str = "Library") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _FILENAME_SAFE.sub(" ", sheet_title).strip()[:31] or "Library"
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")
    widths = [len(header) for header in headers]
    for row in rows:
        sheet.append([value.replace("\n", ", ") for value in row])
        for index, value in enumerate(row):
            widths[index] = max(widths[index], min(len(value), 60))
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = max(8, width + 2)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def render_pdf(
    headers: list[str],
    rows: list[list[str]],
    *,
    weights: list[float],
    title: str = DEFAULT_EXPORT_TITLE,
    generated_label: str = "",
) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        LongTable,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        TableStyle,
    )

    page_size = landscape(A4)
    margin = 12 * mm
    available = page_size[0] - (2 * margin)
    total_weight = sum(weights) or float(len(headers))
    col_widths = [available * (weight / total_weight) for weight in weights]

    header_style = ParagraphStyle(
        "DvExportHeader",
        fontName="Helvetica-Bold",
        fontSize=7.5,
        leading=9,
        textColor=colors.white,
        alignment=TA_LEFT,
    )
    cell_style = ParagraphStyle(
        "DvExportCell",
        fontName="Helvetica",
        fontSize=7,
        leading=8.5,
        alignment=TA_LEFT,
    )
    title_style = ParagraphStyle("DvExportTitle", fontName="Helvetica-Bold", fontSize=13, leading=16)
    meta_style = ParagraphStyle("DvExportMeta", fontName="Helvetica", fontSize=8, leading=10, textColor=colors.grey)

    data: list[list[Any]] = [[Paragraph(_pdf_text(header), header_style) for header in headers]]
    for row in rows:
        data.append([Paragraph(_pdf_text(value), cell_style) for value in row])

    table = LongTable(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2a37")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d0d5dd")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f7fa")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 2.5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
            ]
        )
    )

    stream = io.BytesIO()
    document = SimpleDocTemplate(
        stream,
        pagesize=page_size,
        leftMargin=margin,
        rightMargin=margin,
        topMargin=margin,
        bottomMargin=margin,
        title=title or DEFAULT_EXPORT_TITLE,
        author="DiscVault",
    )
    story: list[Any] = [Paragraph(_pdf_text(title or DEFAULT_EXPORT_TITLE), title_style)]
    if generated_label:
        story.append(Paragraph(_pdf_text(generated_label), meta_style))
    story.append(Spacer(1, 4 * mm))
    story.append(table)
    document.build(story, onFirstPage=_pdf_page_number, onLaterPages=_pdf_page_number)
    return stream.getvalue()


def _pdf_page_number(canvas, document) -> None:  # pragma: no cover - drawing callback
    canvas.saveState()
    canvas.setFont("Helvetica", 7)
    canvas.setFillGray(0.45)
    canvas.drawRightString(document.pagesize[0] - 12, 8, str(canvas.getPageNumber()))
    canvas.restoreState()


def _pdf_text(value: str) -> str:
    escaped = (
        str(value or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )
    return escaped.replace("\n", "<br/>")


def build_export(payload: dict[str, Any], custom_fields: Any = None) -> tuple[bytes, str, str]:
    """Render ``payload`` and return ``(body, filename, mimetype)``."""
    if not isinstance(payload, dict):
        raise NextApiError("Invalid export payload", 400)
    fmt = normalize_format(payload.get("format"))
    columns = normalize_columns(payload.get("columns"), custom_fields)
    rows = normalize_rows(payload.get("rows"), columns)
    if not rows:
        raise NextApiError("There is nothing to export", 400)
    headers = column_headers(columns, payload.get("labels"))
    filename = export_filename(fmt, payload.get("filename"))
    title = str(payload.get("title") or "").strip() or DEFAULT_EXPORT_TITLE

    if fmt == "csv":
        body = render_csv(headers, rows)
    elif fmt == "xlsx":
        body = render_xlsx(headers, rows, sheet_title=title)
    else:
        if len(rows) > MAX_PDF_EXPORT_ROWS:
            raise NextApiError(
                f"A PDF export is limited to {MAX_PDF_EXPORT_ROWS} rows "
                f"({len(rows)} selected). Export to CSV or XLSX instead.",
                400,
            )
        body = render_pdf(
            headers,
            rows,
            weights=column_weights(columns),
            title=title,
            generated_label=str(payload.get("subtitle") or "").strip(),
        )
    return body, filename, EXPORT_MIME_TYPES[fmt]


def register_next_export_routes(flask_app: Flask, *, connect) -> None:  # pragma: no cover - Flask integration
    @flask_app.get("/api/next/collection/export/columns")
    def collection_export_columns():
        # The catalogue is instance-dependent now: two installations
        # legitimately offer different columns, because they define different
        # fields. The built-in half is identical everywhere and is the only half
        # the cross-platform default selection draws from.
        app = _next_app()
        with connect() as conn:
            custom_fields = app.all_custom_field_entities(conn)
        return response({"status": "ok", "columns": export_column_catalogue(custom_fields)})

    @flask_app.post("/api/next/collection/export")
    def collection_export():
        app = _next_app()
        with connect() as conn:
            auth_enabled = app.next_auth_effective_enabled(conn, table_exists)
            user = app.next_auth_current_user(conn) if auth_enabled else None
            if auth_enabled and not user:
                raise NextApiError("Authentication required", 401)
            # On the connection that is already open. The definitions decide
            # which `custom:` column keys are legal, so they have to be read
            # before the payload is validated.
            custom_fields = app.all_custom_field_entities(conn)
        body, filename, mimetype = build_export(
            request.get_json(silent=True, force=True) or {}, custom_fields
        )
        result = flask_app.response_class(body, mimetype=mimetype)
        result.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
        result.headers["Content-Length"] = str(len(body))
        result.headers["Cache-Control"] = "no-store"
        return result
