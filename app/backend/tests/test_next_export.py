import io
import json
import os
import sys
import unittest
from datetime import datetime, timezone


BACKEND_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
REPO_APP_DIR = os.path.abspath(os.path.join(BACKEND_DIR, ".."))
NEXT_VIEWS_UI_PATH = os.path.join(BACKEND_DIR, "next_views_ui.py")
LIBRARY_EXPORT_JS_PATH = os.path.join(REPO_APP_DIR, "frontend", "js", "library-export.js")
I18N_DIR = os.path.join(REPO_APP_DIR, "frontend", "i18n", "next")

if BACKEND_DIR not in sys.path:
    sys.path.insert(0, BACKEND_DIR)

import next_export  # noqa: E402
import next_export_columns  # noqa: E402
import next_static  # noqa: E402
from next_common import NextApiError  # noqa: E402


SAMPLE_ROWS = [
    {
        "title": "Blade Runner",
        "year": "1982",
        "barcode": "5051888210437",
        "format": "4K UHD, Blu-ray",
        "director": "Ridley Scott",
        "actors": "Harrison Ford, Rutger Hauer",
        "studio": "Warner Bros.",
        "contentRating": "16",
        "tags": "Sci-fi",
        "watchActivity": "Watched | 2024-01-05",
    },
    {
        "title": "Amélie",
        "year": "2001",
        "barcode": "8712626049524",
        "format": "Blu-ray",
        "director": "Jean-Pierre Jeunet",
        "actors": "Audrey Tautou",
        "studio": "UGC",
        "contentRating": "12",
        "tags": "Comedy, Romance",
        "watchActivity": "Watchlist",
    },
]


class ExportColumnCatalogueTests(unittest.TestCase):
    def test_catalogue_contains_the_list_view_columns_plus_barcode(self):
        keys = [column["key"] for column in next_export_columns.export_column_catalogue()]
        self.assertEqual(
            keys,
            [
                "title",
                "year",
                "barcode",
                "format",
                "director",
                "actors",
                "studio",
                "contentRating",
                "tags",
                "watchActivity",
            ],
        )

    def test_catalogue_entries_expose_the_picker_contract(self):
        for column in next_export_columns.export_column_catalogue():
            self.assertEqual(
                sorted(column.keys()), ["default", "i18nKey", "key", "label"]
            )
            self.assertIsInstance(column["default"], bool)

    def test_every_column_i18n_key_exists_in_every_locale(self):
        i18n_keys = {column["i18nKey"] for column in next_export_columns.EXPORT_COLUMNS}
        for name in sorted(os.listdir(I18N_DIR)):
            if not name.endswith(".json"):
                continue
            with open(os.path.join(I18N_DIR, name), encoding="utf-8") as handle:
                messages = json.load(handle)
            missing = sorted(key for key in i18n_keys if key not in messages)
            self.assertEqual(missing, [], f"{name} is missing {missing}")


class NormalizeColumnsTests(unittest.TestCase):
    def test_missing_columns_fall_back_to_the_defaults(self):
        self.assertEqual(
            next_export_columns.normalize_columns(None),
            list(next_export_columns.DEFAULT_EXPORT_COLUMN_KEYS),
        )

    def test_columns_are_returned_in_catalogue_order(self):
        self.assertEqual(
            next_export_columns.normalize_columns(["tags", "title", "barcode"]),
            ["title", "barcode", "tags"],
        )

    def test_duplicates_are_collapsed(self):
        self.assertEqual(
            next_export_columns.normalize_columns(["title", "title"]), ["title"]
        )

    def test_unknown_column_is_rejected(self):
        with self.assertRaises(NextApiError) as ctx:
            next_export_columns.normalize_columns(["title", "runtime"])
        self.assertEqual(ctx.exception.status_code, 400)

    def test_empty_selection_is_rejected(self):
        with self.assertRaises(NextApiError):
            next_export_columns.normalize_columns([])

    def test_non_list_is_rejected(self):
        with self.assertRaises(NextApiError):
            next_export_columns.normalize_columns("title")


class NormalizeRowsTests(unittest.TestCase):
    def test_dict_rows_are_mapped_onto_the_selected_columns(self):
        rows = next_export_columns.normalize_rows(SAMPLE_ROWS, ["title", "barcode"])
        self.assertEqual(rows, [["Blade Runner", "5051888210437"], ["Amélie", "8712626049524"]])

    def test_list_rows_are_padded_and_truncated(self):
        rows = next_export_columns.normalize_rows([["a"], ["a", "b", "c"]], ["title", "year"])
        self.assertEqual(rows, [["a", ""], ["a", "b"]])

    def test_missing_values_become_empty_strings(self):
        rows = next_export_columns.normalize_rows([{"title": "X"}], ["title", "barcode"])
        self.assertEqual(rows, [["X", ""]])

    def test_lists_are_joined_and_control_characters_stripped(self):
        rows = next_export_columns.normalize_rows(
            [{"tags": ["Sci-fi", "Cult"], "title": "A\x07B"}], ["title", "tags"]
        )
        self.assertEqual(rows, [["AB", "Sci-fi, Cult"]])

    def test_invalid_row_shape_is_rejected(self):
        with self.assertRaises(NextApiError):
            next_export_columns.normalize_rows([42], ["title"])

    def test_too_many_rows_are_rejected(self):
        oversized = [{"title": "x"}] * (next_export_columns.MAX_EXPORT_ROWS + 1)
        with self.assertRaises(NextApiError):
            next_export_columns.normalize_rows(oversized, ["title"])


class ColumnHeaderTests(unittest.TestCase):
    def test_client_labels_win_over_the_english_fallback(self):
        headers = next_export_columns.column_headers(
            ["title", "barcode"], {"title": "Titel"}
        )
        self.assertEqual(headers, ["Titel", "Barcode"])

    def test_blank_labels_fall_back(self):
        headers = next_export_columns.column_headers(["title"], {"title": "  "})
        self.assertEqual(headers, ["Title"])

    def test_weights_follow_the_catalogue(self):
        weights = next_export_columns.column_weights(["title", "year"])
        self.assertEqual(len(weights), 2)
        self.assertGreater(weights[0], weights[1])


class ExportFilenameTests(unittest.TestCase):
    def test_default_filename_uses_the_current_date(self):
        moment = datetime(2026, 3, 9, tzinfo=timezone.utc)
        self.assertEqual(
            next_export.export_filename("csv", None, now=moment),
            "discvault-library-20260309.csv",
        )

    def test_unsafe_characters_are_replaced(self):
        self.assertEqual(
            next_export.export_filename("xlsx", "../../etc/passwd"),
            "etc-passwd.xlsx",
        )

    def test_existing_extension_is_not_duplicated(self):
        self.assertEqual(next_export.export_filename("pdf", "my-list.pdf"), "my-list.pdf")

    def test_stem_is_capped(self):
        name = next_export.export_filename("csv", "a" * 400)
        self.assertEqual(len(name), 124)


class NormalizeFormatTests(unittest.TestCase):
    def test_default_format_is_csv(self):
        self.assertEqual(next_export.normalize_format(None), "csv")

    def test_format_is_case_insensitive(self):
        self.assertEqual(next_export.normalize_format("XLSX"), "xlsx")

    def test_unsupported_format_is_rejected(self):
        with self.assertRaises(NextApiError) as ctx:
            next_export.normalize_format("docx")
        self.assertEqual(ctx.exception.status_code, 400)


class BuildExportTests(unittest.TestCase):
    def _payload(self, fmt, **overrides):
        payload = {
            "format": fmt,
            "columns": ["title", "year", "barcode", "watchActivity"],
            "rows": SAMPLE_ROWS,
            "labels": {"title": "Titel", "barcode": "Streepjescode"},
            "title": "DiscVault library",
            "subtitle": "2 movies",
        }
        payload.update(overrides)
        return payload

    def test_csv_contains_the_headers_rows_and_barcode_column(self):
        body, filename, mimetype = next_export.build_export(self._payload("csv"))
        self.assertTrue(body.startswith(b"\xef\xbb\xbf"))
        self.assertTrue(filename.endswith(".csv"))
        self.assertEqual(mimetype, next_export.EXPORT_MIME_TYPES["csv"])
        text = body.decode("utf-8-sig")
        lines = text.strip().splitlines()
        self.assertEqual(lines[0], "Titel;Year;Streepjescode;Viewing activity")
        self.assertTrue(lines[1].startswith("Blade Runner;1982;5051888210437;"))
        self.assertIn("8712626049524", lines[2])

    def test_csv_keeps_the_submitted_row_order(self):
        payload = self._payload("csv", rows=list(reversed(SAMPLE_ROWS)))
        text = next_export.build_export(payload)[0].decode("utf-8-sig")
        lines = text.strip().splitlines()
        self.assertTrue(lines[1].startswith("Amélie;"))
        self.assertTrue(lines[2].startswith("Blade Runner;"))

    def test_csv_flattens_newlines(self):
        payload = self._payload("csv", rows=[{"title": "A\nB"}], columns=["title"])
        text = next_export.build_export(payload)[0].decode("utf-8-sig")
        self.assertIn("A / B", text)

    def test_xlsx_is_a_zip_container_with_the_expected_grid(self):
        body, filename, mimetype = next_export.build_export(self._payload("xlsx"))
        self.assertTrue(body.startswith(b"PK\x03\x04"))
        self.assertTrue(filename.endswith(".xlsx"))
        self.assertEqual(mimetype, next_export.EXPORT_MIME_TYPES["xlsx"])
        from openpyxl import load_workbook

        sheet = load_workbook(io.BytesIO(body)).active
        self.assertEqual(
            [cell.value for cell in sheet[1]],
            ["Titel", "Year", "Streepjescode", "Viewing activity"],
        )
        self.assertEqual(sheet.max_row, len(SAMPLE_ROWS) + 1)
        self.assertEqual(sheet["C2"].value, "5051888210437")
        self.assertEqual(sheet.freeze_panes, "A2")

    def test_pdf_is_rendered_in_landscape(self):
        body, filename, mimetype = next_export.build_export(self._payload("pdf"))
        self.assertTrue(body.startswith(b"%PDF"))
        self.assertTrue(filename.endswith(".pdf"))
        self.assertEqual(mimetype, next_export.EXPORT_MIME_TYPES["pdf"])
        self.assertIn(b"/MediaBox [ 0 0 841", body)

    def test_pdf_paginates_a_large_selection(self):
        rows = [dict(SAMPLE_ROWS[0], title=f"Movie {index}") for index in range(400)]
        body = next_export.build_export(self._payload("pdf", rows=rows))[0]
        self.assertTrue(body.startswith(b"%PDF"))
        self.assertGreater(body.count(b"/Type /Page\n"), 1)

    def test_pdf_escapes_markup_in_cell_values(self):
        payload = self._payload("pdf", rows=[{"title": "Fish <&> Chips"}], columns=["title"])
        self.assertTrue(next_export.build_export(payload)[0].startswith(b"%PDF"))

    def test_empty_selection_is_rejected(self):
        with self.assertRaises(NextApiError) as ctx:
            next_export.build_export(self._payload("csv", rows=[]))
        self.assertEqual(ctx.exception.status_code, 400)

    def test_invalid_payload_is_rejected(self):
        with self.assertRaises(NextApiError):
            next_export.build_export("nope")


class ExportWiringTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        with open(NEXT_VIEWS_UI_PATH, encoding="utf-8") as handle:
            cls.app_source = handle.read()
        with open(LIBRARY_EXPORT_JS_PATH, encoding="utf-8") as handle:
            cls.export_js = handle.read()

    def test_export_script_is_whitelisted(self):
        self.assertIn("library-export.js", next_static.NEXT_SCRIPT_ASSETS)

    def test_export_script_is_loaded_by_the_spa(self):
        self.assertIn(
            '<script src="/api/next/app/js/library-export.js" defer></script>',
            self.app_source,
        )

    def test_export_button_lives_in_the_library_toolbar_only(self):
        self.assertEqual(self.app_source.count('id="libraryExportButton"'), 1)

    def test_bridge_exposes_the_export_hooks(self):
        self.assertIn("getExportRows:", self.app_source)
        self.assertIn("getExportColumnLabels:", self.app_source)

    def test_export_row_helper_covers_every_catalogue_column(self):
        start = self.app_source.index("function libraryExportRow(")
        end = self.app_source.index("\n    function ", start + 1)
        body = self.app_source[start:end]
        for key in next_export_columns.EXPORT_COLUMN_KEYS:
            self.assertIn(f"{key}:", body)

    def test_frontend_module_targets_the_export_endpoints(self):
        self.assertIn("/api/next/collection/export/columns", self.export_js)
        self.assertIn("/api/next/collection/export", self.export_js)

    def test_frontend_module_offers_every_export_format(self):
        for fmt in next_export.EXPORT_FORMATS:
            self.assertIn(f'"{fmt}"', self.export_js)

    def test_export_i18n_keys_exist_in_every_locale(self):
        required = [
            "collection.exportLibrary",
            "collection.exportTitle",
            "collection.exportSummary",
            "collection.exportFormat",
            "collection.exportColumns",
            "collection.exportDownload",
            "collection.exportNoColumns",
            "collection.exportEmpty",
            "collection.exportFailed",
            "collection.exportRowCount",
            "collection.exportDocumentTitle",
        ]
        for name in sorted(os.listdir(I18N_DIR)):
            if not name.endswith(".json"):
                continue
            with open(os.path.join(I18N_DIR, name), encoding="utf-8") as handle:
                messages = json.load(handle)
            missing = sorted(key for key in required if key not in messages)
            self.assertEqual(missing, [], f"{name} is missing {missing}")
            for key in ("collection.exportSummary", "collection.exportRowCount"):
                self.assertIn("{count}", messages[key], f"{name}:{key} lost its placeholder")


if __name__ == "__main__":
    unittest.main()
